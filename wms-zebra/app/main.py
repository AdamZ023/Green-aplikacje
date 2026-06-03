from io import BytesIO
import csv
import json
import re
from uuid import uuid4

from fastapi import Depends, FastAPI, HTTPException, Query, Request
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
import qrcode
import qrcode.image.svg
from sqlalchemy import case, func, inspect, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.config import settings
from app.database import Base, engine, get_db
from app.models import (
    AllocationEvent,
    AllocationPlanItem,
    AllocationWorkspace,
    DeliveryImport,
    DeliveryPallet,
    DeliveryPalletContent,
    Item,
    Operation,
    PickingBatch,
    PickingTask,
    ScannerDevice,
    ScannerLoginSession,
    Stock,
)
from app.schemas import (
    AllocationActionOut,
    AllocationContentOut,
    AllocationDeliveryOut,
    AllocationEventOut,
    AllocationEventUndoRequest,
    AllocationImportOut,
    AllocationPalletOut,
    AllocationPlanItemOut,
    AllocationSectionMoveRequest,
    AllocationSectionRequest,
    AllocationWorkspaceCreate,
    AllocationWorkspaceDeleteRequest,
    AllocationWorkspaceOut,
    ItemCreate,
    ItemOut,
    IssueRequest,
    MoveRequest,
    OperationOut,
    PickingBatchOut,
    PickingCancelRequest,
    PickingCompleteRequest,
    PickingFinishRequest,
    PickingImportOut,
    PickingNextRequest,
    PickingTaskOut,
    ReceiveRequest,
    ScannerRegistrationOut,
    ScannerRegistrationRequest,
    ShippingOut,
    StockOut,
    WarehouseStockOut,
)
from app.security import require_api_key
from app.services import (
    WmsError,
    complete_picking_task,
    create_item,
    issue_stock,
    move_stock,
    receive_stock,
    scan_timestamp,
)

APP_VERSION = "20260603-5"
WAREHOUSE_CODE = "9201D"
PICKING_HEADER_ALIASES = {
    "code": {"ean", "barcode", "kod", "kod kreskowy", "sku", "indeks", "index"},
    "quantity": {"ilosc", "qty", "quantity", "szt", "sztuki"},
    "target": {
        "do lokalizacji",
        "lokalizacja docelowa",
        "docelowa lokalizacja",
        "lokalizacja",
        "dest",
        "destination",
        "gdzie",
    },
}
CACHE_HEADERS = {
    "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
    "Pragma": "no-cache",
    "Expires": "0",
}
RESERVED_PICKING_STATUSES = ("pending", "assigned")

Base.metadata.create_all(bind=engine)
inspector = inspect(engine)
if "scanner_devices" not in inspector.get_table_names():
    ScannerDevice.__table__.create(bind=engine)
else:
    scanner_columns = {column["name"] for column in inspector.get_columns("scanner_devices")}
    if "device_uid" not in scanner_columns:
        with engine.begin() as connection:
            connection.execute(text("ALTER TABLE scanner_devices ADD COLUMN device_uid VARCHAR(120)"))
            connection.execute(
                text("CREATE UNIQUE INDEX IF NOT EXISTS ix_scanner_devices_device_uid ON scanner_devices (device_uid)")
            )

if "picking_tasks" not in inspector.get_table_names():
    PickingTask.__table__.create(bind=engine)
else:
    picking_columns = {column["name"] for column in inspector.get_columns("picking_tasks")}
    if "assigned_at" not in picking_columns:
        assigned_at_type = "TIMESTAMP WITH TIME ZONE" if engine.url.get_backend_name() == "postgresql" else "TIMESTAMP"
        with engine.begin() as connection:
            connection.execute(text(f"ALTER TABLE picking_tasks ADD COLUMN assigned_at {assigned_at_type}"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS ix_picking_tasks_assigned_at ON picking_tasks (assigned_at)"))

if "picking_batches" not in inspector.get_table_names():
    PickingBatch.__table__.create(bind=engine)

if "delivery_pallets" not in inspector.get_table_names():
    DeliveryPallet.__table__.create(bind=engine)
else:
    delivery_pallet_columns = {column["name"] for column in inspector.get_columns("delivery_pallets")}
    with engine.begin() as connection:
        if "layout_row" not in delivery_pallet_columns:
            connection.execute(text("ALTER TABLE delivery_pallets ADD COLUMN layout_row VARCHAR(40)"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS ix_delivery_pallets_layout_row ON delivery_pallets (layout_row)"))
        if "layout_position" not in delivery_pallet_columns:
            connection.execute(text("ALTER TABLE delivery_pallets ADD COLUMN layout_position VARCHAR(40)"))
            connection.execute(
                text("CREATE INDEX IF NOT EXISTS ix_delivery_pallets_layout_position ON delivery_pallets (layout_position)")
            )

if "allocation_events" in inspector.get_table_names():
    allocation_event_columns = {column["name"] for column in inspector.get_columns("allocation_events")}
    with engine.begin() as connection:
        if "undo_payload" not in allocation_event_columns:
            connection.execute(text("ALTER TABLE allocation_events ADD COLUMN undo_payload TEXT"))
        if "undone_at" not in allocation_event_columns:
            undone_at_type = "TIMESTAMP WITH TIME ZONE" if engine.url.get_backend_name() == "postgresql" else "TIMESTAMP"
            connection.execute(text(f"ALTER TABLE allocation_events ADD COLUMN undone_at {undone_at_type}"))

app = FastAPI(title="WMS Zebra API", version="0.1.0")
app.mount("/static", StaticFiles(directory="app/static"), name="static")


@app.get("/", include_in_schema=False)
def dashboard() -> FileResponse:
    return FileResponse("app/static/dashboard.html", headers=CACHE_HEADERS)


@app.get("/scanner", include_in_schema=False)
def scanner() -> FileResponse:
    return FileResponse("app/static/scanner.html", headers=CACHE_HEADERS)


@app.get("/zebra-v2", include_in_schema=False)
def zebra_v2() -> FileResponse:
    return FileResponse("app/static/scanner.html", headers=CACHE_HEADERS)


@app.get("/scanner-qr.svg", include_in_schema=False)
def scanner_qr(request: Request) -> Response:
    factory = qrcode.image.svg.SvgPathImage
    session_id = uuid4().hex
    api_key_param = f"&key={settings.wms_api_key}&session={session_id}"
    if settings.wms_public_url:
        scanner_url = f"{settings.wms_public_url.rstrip('/')}/zebra-v2?v={APP_VERSION}{api_key_param}"
    else:
        scanner_url = f"{request.url_for('zebra_v2')}?v={APP_VERSION}{api_key_param}"
    image = qrcode.make(scanner_url, image_factory=factory, box_size=12, border=2)
    buffer = BytesIO()
    image.save(buffer)
    return Response(content=buffer.getvalue(), media_type="image/svg+xml", headers=CACHE_HEADERS)


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/health/db")
def health_db() -> dict[str, str]:
    return {
        "database": engine.url.get_backend_name(),
        "driver": engine.url.get_driver_name(),
    }


@app.post(
    "/api/scanners/register",
    response_model=ScannerRegistrationOut,
    dependencies=[Depends(require_api_key)],
)
def register_scanner(
    payload: ScannerRegistrationRequest | None = None,
    db: Session = Depends(get_db),
) -> ScannerRegistrationOut:
    payload = payload or ScannerRegistrationRequest()
    device_uid = payload.device_uid.strip() if payload.device_uid else None
    session_id = payload.session_id.strip() if payload.session_id else None

    if session_id and device_uid:
        existing_session = db.scalar(
            select(ScannerLoginSession).where(
                ScannerLoginSession.session_id == session_id,
                ScannerLoginSession.device_uid == device_uid,
            )
        )
        if not payload.force_new:
            if existing_session:
                return ScannerRegistrationOut(scanner_id=existing_session.scanner_id)
        elif existing_session:
            db.delete(existing_session)
            db.commit()

        for _ in range(20):
            session_ids = set(
                db.scalars(
                    select(ScannerLoginSession.scanner_id).where(ScannerLoginSession.session_id == session_id)
                )
            )
            next_number = 1
            while True:
                scanner_id = f"ZEBRA-{next_number:02d}"
                if scanner_id not in session_ids:
                    break
                next_number += 1

            db.add(ScannerLoginSession(session_id=session_id, device_uid=device_uid, scanner_id=scanner_id))
            try:
                db.commit()
            except IntegrityError:
                db.rollback()
                continue
            return ScannerRegistrationOut(scanner_id=scanner_id)
        raise HTTPException(status_code=409, detail="Nie mozna nadac ID skanera. Sprobuj ponownie.")

    if device_uid and not payload.force_new:
        existing_device = db.scalar(select(ScannerDevice).where(ScannerDevice.device_uid == device_uid))
        if existing_device:
            return ScannerRegistrationOut(scanner_id=existing_device.scanner_id)

    existing_ids = set(
        db.scalars(select(ScannerDevice.scanner_id).where(ScannerDevice.device_uid.is_not(None)))
    )
    existing_ids.update(db.scalars(select(Operation.scanner_id).where(Operation.scanner_id.like("ZEBRA-%"))))
    existing_numbers = []
    for scanner_id in existing_ids:
        match = re.fullmatch(r"ZEBRA-(\d+)", scanner_id or "")
        if match:
            existing_numbers.append(int(match.group(1)))
    next_number = (max(existing_numbers) if existing_numbers else 0) + 1
    while True:
        scanner_id = f"ZEBRA-{next_number:02d}"
        if scanner_id not in existing_ids:
            break
        next_number += 1

    existing_device = db.scalar(select(ScannerDevice).where(ScannerDevice.device_uid == device_uid)) if device_uid else None
    placeholder = db.scalar(select(ScannerDevice).where(ScannerDevice.scanner_id == scanner_id))
    if placeholder and placeholder.device_uid is None and device_uid:
        db.delete(placeholder)
        db.flush()

    if existing_device:
        existing_device.scanner_id = scanner_id
    else:
        db.add(ScannerDevice(scanner_id=scanner_id, device_uid=device_uid))
    db.commit()
    return ScannerRegistrationOut(scanner_id=scanner_id)


@app.get("/api/items", response_model=list[ItemOut], dependencies=[Depends(require_api_key)])
def list_items(db: Session = Depends(get_db)) -> list[Item]:
    return list(db.scalars(select(Item).order_by(Item.sku)))


@app.post("/api/items", response_model=ItemOut, dependencies=[Depends(require_api_key)])
def add_item(payload: ItemCreate, db: Session = Depends(get_db)) -> Item:
    try:
        return create_item(db, payload.sku, payload.name, payload.barcode)
    except WmsError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/stock", response_model=list[StockOut], dependencies=[Depends(require_api_key)])
def list_stock(db: Session = Depends(get_db)) -> list[StockOut]:
    reserved_by_sku_location = get_reserved_by_sku_location(db)
    excluded_locations = get_picking_target_locations(db)
    filters = [Stock.quantity > 0]
    if excluded_locations:
        filters.append(Stock.location.not_in(excluded_locations))
    rows = db.execute(
        select(Item.sku, Item.barcode, Item.name, Stock.location, Stock.quantity)
        .join(Stock, Stock.item_id == Item.id)
        .where(*filters)
        .order_by(Item.sku, Stock.location)
    ).all()
    stock_rows = []
    for sku, barcode, name, location, quantity in rows:
        latest_operation = db.scalar(
            select(Operation)
            .where(
                Operation.sku == sku,
                (
                    (Operation.to_location == location)
                    | (Operation.from_location == location)
                ),
            )
            .order_by(Operation.id.desc())
            .limit(1)
        )
        stock_rows.append(
            StockOut(
                sku=sku,
                barcode=barcode,
                name=name,
                location=location,
                quantity=quantity,
                reserved_quantity=reserved_by_sku_location.get((sku, location), 0),
                operator=latest_operation.operator if latest_operation else None,
                scanner_id=latest_operation.scanner_id if latest_operation else None,
                scan_at=latest_operation.created_at if latest_operation else None,
            )
        )
    return stock_rows


@app.get("/api/warehouse-stock", response_model=list[WarehouseStockOut], dependencies=[Depends(require_api_key)])
def list_warehouse_stock(db: Session = Depends(get_db)) -> list[WarehouseStockOut]:
    reserved_by_sku = get_reserved_by_sku(db)
    excluded_locations = get_picking_target_locations(db)
    stock_join_filter = Stock.item_id == Item.id
    if excluded_locations:
        stock_join_filter = stock_join_filter & Stock.location.not_in(excluded_locations)
    rows = db.execute(
        select(
            Item.sku,
            Item.barcode,
            Item.name,
            func.coalesce(func.sum(Stock.quantity), 0),
        )
        .outerjoin(Stock, stock_join_filter)
        .group_by(Item.sku, Item.barcode, Item.name)
        .order_by(Item.sku)
    ).all()
    stock_rows = []
    for sku, barcode, name, quantity in rows:
        latest_operation = db.scalar(
            select(Operation)
            .where(Operation.sku == sku)
            .order_by(Operation.id.desc())
            .limit(1)
        )
        stock_rows.append(
            WarehouseStockOut(
                sku=sku,
                barcode=barcode,
                name=name,
                warehouse=WAREHOUSE_CODE,
                quantity=quantity,
                reserved_quantity=reserved_by_sku.get(sku, 0),
                operator=latest_operation.operator if latest_operation else None,
                scanner_id=latest_operation.scanner_id if latest_operation else None,
                scan_at=latest_operation.created_at if latest_operation else None,
            )
        )
    return stock_rows


@app.post("/api/stock/receive", response_model=StockOut, dependencies=[Depends(require_api_key)])
def receive(payload: ReceiveRequest, db: Session = Depends(get_db)) -> StockOut:
    try:
        stock = receive_stock(db, payload.sku, payload.location, payload.quantity, payload.scanner_id, payload.operator)
    except WmsError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return StockOut(
        sku=stock.item.sku,
        barcode=stock.item.barcode,
        name=stock.item.name,
        location=stock.location,
        quantity=stock.quantity,
        operator=payload.operator,
        scanner_id=payload.scanner_id,
        scan_at=None,
    )


@app.post("/api/stock/issue", response_model=StockOut, dependencies=[Depends(require_api_key)])
def issue(payload: IssueRequest, db: Session = Depends(get_db)) -> StockOut:
    try:
        stock = issue_stock(db, payload.sku, payload.location, payload.quantity, payload.scanner_id, payload.operator)
    except WmsError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return StockOut(
        sku=stock.item.sku,
        barcode=stock.item.barcode,
        name=stock.item.name,
        location=stock.location,
        quantity=stock.quantity,
        operator=payload.operator,
        scanner_id=payload.scanner_id,
        scan_at=None,
    )


@app.post("/api/stock/move", response_model=list[StockOut], dependencies=[Depends(require_api_key)])
def move(payload: MoveRequest, db: Session = Depends(get_db)) -> list[StockOut]:
    try:
        source, target = move_stock(
            db,
            payload.sku,
            payload.from_location,
            payload.to_location,
            payload.quantity,
            payload.scanner_id,
            payload.operator,
        )
    except WmsError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return [
        StockOut(
            sku=source.item.sku,
            barcode=source.item.barcode,
            name=source.item.name,
            location=source.location,
            quantity=source.quantity,
            operator=payload.operator,
            scanner_id=payload.scanner_id,
            scan_at=None,
        ),
        StockOut(
            sku=target.item.sku,
            barcode=target.item.barcode,
            name=target.item.name,
            location=target.location,
            quantity=target.quantity,
            operator=payload.operator,
            scanner_id=payload.scanner_id,
            scan_at=None,
        ),
    ]


@app.post("/api/picking/import", response_model=PickingImportOut, dependencies=[Depends(require_api_key)])
async def import_picking(request: Request, db: Session = Depends(get_db)) -> PickingImportOut:
    filename = request.headers.get("X-Filename", "picking.csv")
    content = await request.body()
    try:
        rows = parse_picking_file(filename, content)
        batch_id = uuid4().hex[:12].upper()
        db.add(PickingBatch(batch_id=batch_id, source_filename=filename[:240]))
        created, blocked = create_picking_tasks(db, batch_id, rows)
    except WmsError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return PickingImportOut(batch_id=batch_id, created=created, blocked=blocked)


@app.get("/api/picking/batches", response_model=list[PickingBatchOut], dependencies=[Depends(require_api_key)])
def list_picking_batches(db: Session = Depends(get_db)) -> list[PickingBatchOut]:
    block_invalid_picking_sources(db)
    task_rows = db.execute(
        select(
            PickingTask.batch_id,
            func.count(PickingTask.id),
            func.sum(case((PickingTask.source_location.is_not(None), 1), else_=0)),
            func.sum(case((PickingTask.status == "done", 1), else_=0)),
            func.sum(case((PickingTask.status == "assigned", 1), else_=0)),
            func.sum(case((PickingTask.status == "canceled", 1), else_=0)),
            func.sum(case((PickingTask.status == "closed", 1), else_=0)),
            func.max(PickingTask.created_at),
        )
        .group_by(PickingTask.batch_id)
        .order_by(func.max(PickingTask.created_at).desc())
    ).all()
    batch_ids = [row[0] for row in task_rows]
    batch_by_id = {}
    if batch_ids:
        batch_by_id = {
            batch.batch_id: batch
            for batch in db.scalars(select(PickingBatch).where(PickingBatch.batch_id.in_(batch_ids)))
        }
    return [
        picking_batch_out(
            batch_id=row[0],
            source_filename=batch_by_id[row[0]].source_filename if row[0] in batch_by_id else None,
            total_tasks=int(row[1] or 0),
            assigned_tasks=int(row[2] or 0),
            done_tasks=int(row[3] or 0),
            active_tasks=int(row[4] or 0),
            canceled_tasks=int(row[5] or 0),
            closed_tasks=int(row[6] or 0),
            created_at=batch_by_id[row[0]].created_at if row[0] in batch_by_id else row[7],
        )
        for row in task_rows
    ]


@app.get("/api/picking/tasks", response_model=list[PickingTaskOut], dependencies=[Depends(require_api_key)])
def list_picking_tasks(
    status: str | None = Query(default=None),
    batch_id: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    db: Session = Depends(get_db),
) -> list[PickingTaskOut]:
    block_invalid_picking_sources(db)
    query = select(PickingTask)
    if status:
        query = query.where(PickingTask.status == status)
    if batch_id:
        query = query.where(PickingTask.batch_id == batch_id)
    query = query.order_by(PickingTask.id if batch_id else PickingTask.id.desc()).limit(limit)
    tasks = list(db.scalars(query))
    return [picking_task_out(db, task) for task in tasks]


@app.post("/api/picking/cancel", response_model=PickingBatchOut, dependencies=[Depends(require_api_key)])
def cancel_picking(payload: PickingCancelRequest, db: Session = Depends(get_db)) -> PickingBatchOut:
    tasks = list(
        db.scalars(
            select(PickingTask)
            .where(PickingTask.batch_id == payload.batch_id)
            .with_for_update()
        )
    )
    if not tasks:
        raise HTTPException(status_code=404, detail="Nie znaleziono pickingu.")

    cancelable = [task for task in tasks if task.status != "done"]
    if not cancelable:
        raise HTTPException(status_code=400, detail="Picking jest juz zebrany i nie mozna go anulowac.")

    for task in cancelable:
        task.status = "canceled"
        task.scanner_id = None
        task.operator = None
        task.assigned_at = None
    db.commit()

    batch = db.scalar(select(PickingBatch).where(PickingBatch.batch_id == payload.batch_id))
    return picking_batch_out(
        batch_id=payload.batch_id,
        source_filename=batch.source_filename if batch else None,
        total_tasks=len(tasks),
        assigned_tasks=sum(1 for task in tasks if task.source_location),
        done_tasks=sum(1 for task in tasks if task.status == "done"),
        active_tasks=0,
        canceled_tasks=sum(1 for task in tasks if task.status == "canceled"),
        closed_tasks=sum(1 for task in tasks if task.status == "closed"),
        created_at=batch.created_at if batch else tasks[0].created_at,
    )


@app.post("/api/picking/finish", response_model=PickingBatchOut, dependencies=[Depends(require_api_key)])
def finish_picking(payload: PickingFinishRequest, db: Session = Depends(get_db)) -> PickingBatchOut:
    tasks = list(
        db.scalars(
            select(PickingTask)
            .where(PickingTask.batch_id == payload.batch_id)
            .with_for_update()
        )
    )
    if not tasks:
        raise HTTPException(status_code=404, detail="Nie znaleziono pickingu.")

    done_count = sum(1 for task in tasks if task.status == "done")
    finishable = [task for task in tasks if task.status not in {"done", "closed"}]
    if not finishable and done_count >= len(tasks):
        raise HTTPException(status_code=400, detail="Picking jest juz zebrany.")
    if not finishable and done_count < len(tasks):
        raise HTTPException(status_code=400, detail="Picking jest juz zakonczony czesciowo.")

    for task in finishable:
        task.status = "closed"
        task.scanner_id = None
        task.operator = None
        task.assigned_at = None
    db.commit()

    batch = db.scalar(select(PickingBatch).where(PickingBatch.batch_id == payload.batch_id))
    return picking_batch_out(
        batch_id=payload.batch_id,
        source_filename=batch.source_filename if batch else None,
        total_tasks=len(tasks),
        assigned_tasks=sum(1 for task in tasks if task.source_location),
        done_tasks=sum(1 for task in tasks if task.status == "done"),
        active_tasks=0,
        canceled_tasks=sum(1 for task in tasks if task.status == "canceled"),
        closed_tasks=sum(1 for task in tasks if task.status == "closed"),
        created_at=batch.created_at if batch else tasks[0].created_at,
    )


@app.get("/api/shipping", response_model=list[ShippingOut], dependencies=[Depends(require_api_key)])
def list_shipping(limit: int = Query(default=500, ge=1, le=1000), db: Session = Depends(get_db)) -> list[ShippingOut]:
    shipping_batch_statuses = get_shipping_picking_batch_statuses(db)
    if not shipping_batch_statuses:
        return []

    shipping_batch_ids = list(shipping_batch_statuses)
    batches_by_id = {
        batch.batch_id: batch
        for batch in db.scalars(select(PickingBatch).where(PickingBatch.batch_id.in_(shipping_batch_ids)))
    }
    tasks = list(
        db.scalars(
            select(PickingTask)
            .where(
                PickingTask.batch_id.in_(shipping_batch_ids),
                PickingTask.status == "done",
            )
            .order_by(PickingTask.picked_at.desc(), PickingTask.id.desc())
            .limit(limit)
        )
    )
    return [
        shipping_out(db, task, batches_by_id.get(task.batch_id), shipping_batch_statuses.get(task.batch_id, "zebrany"))
        for task in tasks
    ]


@app.post(
    "/api/allocations/workspaces",
    response_model=AllocationWorkspaceOut,
    dependencies=[Depends(require_api_key)],
)
def create_allocation_workspace(
    payload: AllocationWorkspaceCreate,
    db: Session = Depends(get_db),
) -> AllocationWorkspaceOut:
    workspace = AllocationWorkspace(
        workspace_id=uuid4().hex[:12].upper(),
        name=payload.name.strip(),
        status="robocza",
    )
    db.add(workspace)
    add_allocation_event(
        db,
        workspace.workspace_id,
        "utworzenie_alokacji",
        f"Utworzono alokacje robocza: {workspace.name}.",
    )
    db.commit()
    db.refresh(workspace)
    return allocation_workspace_out(db, workspace)


@app.get(
    "/api/allocations/workspaces",
    response_model=list[AllocationWorkspaceOut],
    dependencies=[Depends(require_api_key)],
)
def list_allocation_workspaces(db: Session = Depends(get_db)) -> list[AllocationWorkspaceOut]:
    workspaces = list(db.scalars(select(AllocationWorkspace).order_by(AllocationWorkspace.id.desc())))
    return [allocation_workspace_out(db, workspace) for workspace in workspaces]


@app.delete(
    "/api/allocations/workspaces",
    response_model=AllocationActionOut,
    dependencies=[Depends(require_api_key)],
)
def delete_allocation_workspace(
    payload: AllocationWorkspaceDeleteRequest,
    db: Session = Depends(get_db),
) -> AllocationActionOut:
    workspace = ensure_allocation_workspace(db, payload.workspace_id)
    add_allocation_event(
        db,
        payload.workspace_id,
        "usuniecie_alokacji",
        f"Usunieto alokacje robocza: {workspace.name}.",
    )
    delete_allocation_workspace_rows(db, payload.workspace_id)
    db.commit()
    return AllocationActionOut(message=f"Usunieto alokacje robocza {payload.workspace_id}.")


@app.post(
    "/api/allocations/delivery-import",
    response_model=AllocationImportOut,
    dependencies=[Depends(require_api_key)],
)
async def import_allocation_delivery(request: Request, db: Session = Depends(get_db)) -> AllocationImportOut:
    workspace_id = request.headers.get("X-Workspace-Id", "").strip()
    filename = request.headers.get("X-Filename", "rozladunek.xlsx").strip()
    if not workspace_id:
        raise HTTPException(status_code=400, detail="Wybierz alokacje robocza.")
    ensure_allocation_workspace(db, workspace_id)
    content = await request.body()
    try:
        result = import_delivery_xlsx(db, workspace_id, filename, content)
    except WmsError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    delivery = db.scalar(
        select(DeliveryImport).where(
            DeliveryImport.workspace_id == workspace_id,
            DeliveryImport.source_filename == result.source_filename,
        )
    )
    add_allocation_event(
        db,
        workspace_id,
        "import_rozladunku",
        result.message,
        source_filename=result.source_filename,
        pallet_count=parse_count_from_message(result.message, "palet"),
        carton_count=parse_carton_count_from_message(result.message),
        undo_payload={"action": "delete_delivery", "delivery_id": delivery.delivery_id} if delivery else None,
    )
    db.commit()
    return result


@app.get(
    "/api/allocations/deliveries",
    response_model=list[AllocationDeliveryOut],
    dependencies=[Depends(require_api_key)],
)
def list_allocation_deliveries(
    workspace_id: str = Query(min_length=1),
    db: Session = Depends(get_db),
) -> list[AllocationDeliveryOut]:
    ensure_allocation_workspace(db, workspace_id)
    pallet_counts = dict(
        db.execute(
            select(DeliveryPallet.delivery_id, func.count(DeliveryPallet.id))
            .where(DeliveryPallet.workspace_id == workspace_id)
            .group_by(DeliveryPallet.delivery_id)
        ).all()
    )
    deliveries = list(
        db.scalars(
            select(DeliveryImport)
            .where(DeliveryImport.workspace_id == workspace_id)
            .order_by(DeliveryImport.created_at.desc(), DeliveryImport.source_filename)
        )
    )
    return [
        AllocationDeliveryOut(
            delivery_id=delivery.delivery_id,
            delivery_ref=delivery.delivery_ref,
            source_filename=delivery.source_filename,
            total_cartons=delivery.total_cartons,
            pallet_count=int(pallet_counts.get(delivery.delivery_id, 0) or 0),
            created_at=delivery.created_at,
        )
        for delivery in deliveries
    ]


@app.delete(
    "/api/allocations/deliveries/{delivery_id}",
    response_model=AllocationActionOut,
    dependencies=[Depends(require_api_key)],
)
def delete_allocation_delivery(delivery_id: str, db: Session = Depends(get_db)) -> AllocationActionOut:
    delivery = db.scalar(select(DeliveryImport).where(DeliveryImport.delivery_id == delivery_id))
    if not delivery:
        raise HTTPException(status_code=404, detail="Nie znaleziono pliku rozladunku w tej alokacji.")
    source_filename = delivery.source_filename
    workspace_id = delivery.workspace_id
    pallet_count = db.scalar(select(func.count(DeliveryPallet.id)).where(DeliveryPallet.delivery_id == delivery_id))
    carton_count = delivery.total_cartons
    add_allocation_event(
        db,
        workspace_id,
        "usuniecie_rozladunku",
        f"Wycofano dostawe z alokacji: {source_filename}.",
        source_filename=source_filename,
        pallet_count=int(pallet_count or 0),
        carton_count=carton_count,
    )
    delete_delivery_import(db, delivery_id)
    db.commit()
    return AllocationActionOut(message=f"Wycofano dostawe z alokacji: {source_filename}.")


@app.post(
    "/api/allocations/plan-import",
    response_model=AllocationImportOut,
    dependencies=[Depends(require_api_key)],
)
async def import_allocation_plan(request: Request, db: Session = Depends(get_db)) -> AllocationImportOut:
    workspace_id = request.headers.get("X-Workspace-Id", "").strip()
    filename = request.headers.get("X-Filename", "alokacja.xlsx").strip()
    if not workspace_id:
        raise HTTPException(status_code=400, detail="Wybierz alokacje robocza.")
    ensure_allocation_workspace(db, workspace_id)
    content = await request.body()
    try:
        result = import_plan_xlsx(db, workspace_id, filename, content)
    except WmsError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    add_allocation_event(
        db,
        workspace_id,
        "import_planu_alokacji",
        result.message,
        source_filename=result.source_filename,
        carton_count=result.imported,
        undo_payload={"action": "delete_plan", "source_filename": result.source_filename},
    )
    db.commit()
    return result


@app.post(
    "/api/allocations/sections/remove",
    response_model=AllocationActionOut,
    dependencies=[Depends(require_api_key)],
)
def remove_allocation_section(
    payload: AllocationSectionRequest,
    db: Session = Depends(get_db),
) -> AllocationActionOut:
    ensure_allocation_workspace(db, payload.workspace_id)
    pallets = get_section_pallets(db, payload.workspace_id, payload.sku, payload.color)
    if not pallets:
        raise HTTPException(status_code=404, detail="Nie znaleziono sekcji MDK w tej alokacji.")
    previous_state = allocation_pallet_snapshot(pallets)
    for pallet in pallets:
        pallet.allocation_status = "usuniete_z_alokacji"
        pallet.layout_position = None
    add_allocation_event(
        db,
        payload.workspace_id,
        "usuniecie_mdk",
        f"Usunieto sekcje z rozstawienia: {payload.sku}{f' / {payload.color}' if payload.color else ''}.",
        sku=payload.sku,
        color=payload.color,
        pallet_count=len(pallets),
        undo_payload={"action": "restore_pallets", "pallets": previous_state},
    )
    db.commit()
    return AllocationActionOut(message=f"Usunieto sekcje z rozstawienia: {payload.sku}.")


@app.post(
    "/api/allocations/sections/move",
    response_model=AllocationActionOut,
    dependencies=[Depends(require_api_key)],
)
def move_allocation_section(
    payload: AllocationSectionMoveRequest,
    db: Session = Depends(get_db),
) -> AllocationActionOut:
    ensure_allocation_workspace(db, payload.workspace_id)
    pallets = get_section_pallets(db, payload.workspace_id, payload.sku, payload.color)
    if not pallets:
        raise HTTPException(status_code=404, detail="Nie znaleziono sekcji MDK w tej alokacji.")
    previous_state = allocation_pallet_snapshot(pallets)
    old_positions = ", ".join(sorted({pallet.layout_position or "" for pallet in pallets if pallet.layout_position}))
    set_section_position(pallets, payload.target_position.strip())
    add_allocation_event(
        db,
        payload.workspace_id,
        "przeniesienie_mdk",
        f"Przeniesiono sekcje {payload.sku}{f' / {payload.color}' if payload.color else ''}: {old_positions or '-'} -> {payload.target_position}.",
        sku=payload.sku,
        color=payload.color,
        pallet_count=len(pallets),
        undo_payload={"action": "restore_pallets", "pallets": previous_state},
    )
    db.commit()
    return AllocationActionOut(message=f"Przeniesiono sekcje {payload.sku} na pozycje {payload.target_position}.")


@app.post(
    "/api/allocations/compact",
    response_model=AllocationActionOut,
    dependencies=[Depends(require_api_key)],
)
def compact_allocation_layout(
    payload: AllocationWorkspaceDeleteRequest,
    db: Session = Depends(get_db),
) -> AllocationActionOut:
    ensure_allocation_workspace(db, payload.workspace_id)
    pallets = list(
        db.scalars(
            select(DeliveryPallet).where(
                DeliveryPallet.workspace_id == payload.workspace_id,
                DeliveryPallet.allocation_status != "usuniete_z_alokacji",
            )
        )
    )
    previous_state = allocation_pallet_snapshot(pallets)
    compact_layout_positions(db, payload.workspace_id)
    add_allocation_event(
        db,
        payload.workspace_id,
        "zsuniecie_palet",
        "Zsunieto palety i usunieto puste miejsca na mapie.",
        pallet_count=len(pallets),
        undo_payload={"action": "restore_pallets", "pallets": previous_state},
    )
    db.commit()
    return AllocationActionOut(message="Zsunieto palety i usunieto puste miejsca na mapie.")


@app.get(
    "/api/allocations/events",
    response_model=list[AllocationEventOut],
    dependencies=[Depends(require_api_key)],
)
def list_allocation_events(
    workspace_id: str = Query(min_length=1),
    limit: int = Query(default=200, ge=1, le=500),
    db: Session = Depends(get_db),
) -> list[AllocationEvent]:
    ensure_allocation_workspace(db, workspace_id)
    return list(
        db.scalars(
            select(AllocationEvent)
            .where(AllocationEvent.workspace_id == workspace_id)
            .order_by(AllocationEvent.created_at.desc(), AllocationEvent.id.desc())
            .limit(limit)
        )
    )


@app.post(
    "/api/allocations/events/undo",
    response_model=AllocationActionOut,
    dependencies=[Depends(require_api_key)],
)
def undo_allocation_event(
    payload: AllocationEventUndoRequest,
    db: Session = Depends(get_db),
) -> AllocationActionOut:
    event = db.scalar(select(AllocationEvent).where(AllocationEvent.id == payload.event_id).with_for_update())
    if not event:
        raise HTTPException(status_code=404, detail="Nie znaleziono operacji w historii alokacji.")
    ensure_allocation_workspace(db, event.workspace_id)
    if not event.undo_payload:
        raise HTTPException(status_code=400, detail="Tej operacji nie da sie cofnac.")
    if event.undone_at is not None:
        raise HTTPException(status_code=400, detail="Ta operacja jest juz cofnieta.")

    try:
        undo_payload = json.loads(event.undo_payload)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Brak poprawnych danych cofania tej operacji.") from exc

    action = undo_payload.get("action")
    if action == "restore_pallets":
        restored = restore_allocation_pallet_snapshot(db, event.workspace_id, undo_payload.get("pallets") or [])
        undo_description = f"Cofnieto operacje: {event.description} Przywrocono palety: {restored}."
    elif action == "delete_delivery":
        delivery_id = undo_payload.get("delivery_id")
        delivery = db.scalar(select(DeliveryImport).where(DeliveryImport.delivery_id == delivery_id)) if delivery_id else None
        if not delivery:
            raise HTTPException(status_code=400, detail="Nie mozna cofnac importu rozladunku, bo dostawa nie istnieje.")
        source_filename = delivery.source_filename
        delete_delivery_import(db, delivery.delivery_id)
        undo_description = f"Cofnieto import rozladunku: {source_filename}."
    elif action == "delete_plan":
        source_filename = undo_payload.get("source_filename")
        if not source_filename:
            raise HTTPException(status_code=400, detail="Brak pliku planu do cofniecia.")
        plan_rows = list(
            db.scalars(
                select(AllocationPlanItem).where(
                    AllocationPlanItem.workspace_id == event.workspace_id,
                    AllocationPlanItem.source_filename == source_filename,
                )
            )
        )
        for row in plan_rows:
            db.delete(row)
        undo_description = f"Cofnieto import planu alokacji: {source_filename}. Usunieto pozycji: {len(plan_rows)}."
    else:
        raise HTTPException(status_code=400, detail="Nieznany typ cofania operacji.")

    event.undone_at = scan_timestamp()
    add_allocation_event(
        db,
        event.workspace_id,
        "cofniecie_operacji",
        undo_description,
        source_filename=event.source_filename,
        sku=event.sku,
        color=event.color,
        pallet_count=event.pallet_count,
        carton_count=event.carton_count,
    )
    db.commit()
    return AllocationActionOut(message=undo_description)


@app.get(
    "/api/allocations/pallets",
    response_model=list[AllocationPalletOut],
    dependencies=[Depends(require_api_key)],
)
def list_allocation_pallets(
    workspace_id: str = Query(min_length=1),
    db: Session = Depends(get_db),
) -> list[AllocationPalletOut]:
    ensure_allocation_workspace(db, workspace_id)
    plan_skus, plan_eans = get_allocation_plan_keys(db, workspace_id)
    imports_by_delivery = get_delivery_imports_by_id(db, workspace_id)
    contents_by_pallet = get_allocation_contents_by_pallet(db, workspace_id)
    layout_by_pallet = build_allocation_layout(db, workspace_id, contents_by_pallet)
    pallets = list(
        db.scalars(
            select(DeliveryPallet)
            .where(
                DeliveryPallet.workspace_id == workspace_id,
                DeliveryPallet.allocation_status != "usuniete_z_alokacji",
            )
            .order_by(DeliveryPallet.delivery_id.desc(), DeliveryPallet.pallet_code)
        )
    )
    rows = []
    for pallet in pallets:
        contents = contents_by_pallet.get(pallet.pallet_code, [])
        status = allocation_content_status(contents, plan_skus, plan_eans)
        delivery_import = imports_by_delivery.get(pallet.delivery_id)
        rows.append(
            AllocationPalletOut(
                pallet_code=pallet.pallet_code,
                pallet_no=pallet.pallet_no,
                delivery_ref=delivery_import.delivery_ref if delivery_import else None,
                source_filename=delivery_import.source_filename if delivery_import else None,
                total_cartons=pallet.total_cartons,
                status=status,
                layout_row=pallet.layout_row or layout_by_pallet.get(pallet.pallet_code, {}).get("row"),
                layout_position=pallet.layout_position or layout_by_pallet.get(pallet.pallet_code, {}).get("position"),
                sku_list=", ".join(sorted({content.sku for content in contents if content.sku})),
                ean_list=", ".join(sorted({content.ean for content in contents if content.ean})),
            )
        )
    return rows


@app.get(
    "/api/allocations/contents",
    response_model=list[AllocationContentOut],
    dependencies=[Depends(require_api_key)],
)
def list_allocation_contents(
    workspace_id: str = Query(min_length=1),
    db: Session = Depends(get_db),
) -> list[AllocationContentOut]:
    ensure_allocation_workspace(db, workspace_id)
    plan_skus, plan_eans = get_allocation_plan_keys(db, workspace_id)
    imports_by_delivery = get_delivery_imports_by_id(db, workspace_id)
    contents_by_pallet = get_allocation_contents_by_pallet(db, workspace_id)
    contents = sorted(
        [content for rows in contents_by_pallet.values() for content in rows],
        key=lambda content: (content.delivery_id, content.pallet_code, content.sku),
        reverse=True,
    )
    rows = []
    for content in contents:
        delivery_import = imports_by_delivery.get(content.delivery_id)
        rows.append(
            AllocationContentOut(
                delivery_ref=delivery_import.delivery_ref if delivery_import else None,
                source_filename=delivery_import.source_filename if delivery_import else None,
                pallet_code=content.pallet_code,
                sku=content.sku,
                color=content.color,
                kind=content.kind,
                size=content.size,
                ean=content.ean,
                quantity_cartons=content.quantity_cartons,
                status=allocation_row_status(content.sku, content.ean, plan_skus, plan_eans),
            )
        )
    return rows


@app.get(
    "/api/allocations/plan",
    response_model=list[AllocationPlanItemOut],
    dependencies=[Depends(require_api_key)],
)
def list_allocation_plan(
    workspace_id: str = Query(min_length=1),
    db: Session = Depends(get_db),
) -> list[AllocationPlanItemOut]:
    ensure_allocation_workspace(db, workspace_id)
    contents_skus, contents_eans = get_allocation_content_keys(db, workspace_id)
    rows = list(
        db.scalars(
            select(AllocationPlanItem)
            .where(AllocationPlanItem.workspace_id == workspace_id)
            .order_by(AllocationPlanItem.id)
        )
    )
    return [
        AllocationPlanItemOut(
            mdk=item.mdk,
            color=item.color,
            supplier=item.supplier,
            delivery_plan=item.delivery_plan,
            ean_prepack=item.ean_prepack,
            source_filename=item.source_filename,
            status="jest w dostawach"
            if allocation_plan_item_found(item, contents_skus, contents_eans)
            else "brak w dostawach",
        )
        for item in rows
    ]


@app.post("/api/picking/next", response_model=PickingTaskOut | None, dependencies=[Depends(require_api_key)])
def next_picking_task(payload: PickingNextRequest, db: Session = Depends(get_db)) -> PickingTaskOut | None:
    block_invalid_picking_sources(db)
    excluded_locations = get_picking_target_locations(db)
    valid_source_filters = [PickingTask.source_location.is_not(None)]
    if excluded_locations:
        valid_source_filters.append(PickingTask.source_location.not_in(excluded_locations))
    existing_task = db.scalar(
        select(PickingTask)
        .where(
            PickingTask.batch_id == payload.batch_id,
            PickingTask.status == "assigned",
            PickingTask.scanner_id == payload.scanner_id,
            *valid_source_filters,
        )
        .order_by(PickingTask.id)
        .limit(1)
    )
    if existing_task:
        return picking_task_out(db, existing_task)

    task = db.scalar(
        select(PickingTask)
        .where(
            PickingTask.batch_id == payload.batch_id,
            PickingTask.status == "pending",
            *valid_source_filters,
        )
        .order_by(PickingTask.id)
        .with_for_update(skip_locked=True)
        .limit(1)
    )
    if task:
        task.status = "assigned"
        task.scanner_id = payload.scanner_id
        task.operator = payload.operator
        task.assigned_at = scan_timestamp()
        db.commit()
        db.refresh(task)
    return picking_task_out(db, task) if task else None


@app.post("/api/picking/complete", response_model=PickingTaskOut, dependencies=[Depends(require_api_key)])
def complete_picking(payload: PickingCompleteRequest, db: Session = Depends(get_db)) -> PickingTaskOut:
    try:
        task = complete_picking_task(
            db,
            payload.task_id,
            payload.sku,
            payload.source_location,
            payload.target_location,
            payload.scanner_id,
            payload.operator,
        )
    except WmsError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return picking_task_out(db, task)


@app.get("/api/operations", response_model=list[OperationOut], dependencies=[Depends(require_api_key)])
def operations(
    limit: int = 100,
    operation_type: str | None = None,
    db: Session = Depends(get_db),
) -> list[Operation]:
    limit = max(1, min(int(limit), 500))
    query = select(Operation)
    if operation_type:
        query = query.where(Operation.operation_type == operation_type)
    return list(db.scalars(query.order_by(Operation.id.desc()).limit(limit)))


def parse_picking_file(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else "csv"
    if suffix == "xlsx":
        return parse_picking_xlsx(content)
    if suffix == "csv":
        return parse_picking_csv(content)
    raise WmsError("Obslugiwane sa tylko pliki CSV i XLSX.")


def parse_picking_csv(content: bytes) -> list[dict[str, str]]:
    for encoding in ("utf-8-sig", "cp1250"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise WmsError("Nie mozna odczytac kodowania pliku CSV.")

    sample = text[:2048]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    return list(csv.DictReader(text.splitlines(), delimiter=delimiter))


def parse_picking_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise WmsError("Brakuje biblioteki openpyxl do odczytu XLSX.") from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    parsed = []
    for row in rows[1:]:
        parsed.append({headers[index]: cell_to_text(value) for index, value in enumerate(row)})
    return parsed


def create_picking_tasks(db: Session, batch_id: str, rows: list[dict[str, str]]) -> tuple[int, int]:
    if not rows:
        raise WmsError("Plik picking jest pusty.")

    created = 0
    blocked = 0
    reserved_by_sku_location = get_reserved_by_sku_location(db)
    excluded_locations = get_picking_target_locations(db)
    for row_number, row in enumerate(rows, start=2):
        normalized = {normalize_header(key): str(value or "").strip() for key, value in row.items()}
        code = normalize_code_value(get_picking_value(normalized, "code"))
        quantity_text = get_picking_value(normalized, "quantity")
        target = get_picking_value(normalized, "target")
        if not code and not quantity_text and not target:
            continue
        if not code or not quantity_text or not target:
            raise WmsError(f"Wiersz {row_number}: wymagane sa produkt, ilosc i lokalizacja docelowa.")
        try:
            quantity = int(float(quantity_text.replace(",", ".")))
        except ValueError as exc:
            raise WmsError(f"Wiersz {row_number}: ilosc musi byc liczba.") from exc
        if quantity < 1:
            raise WmsError(f"Wiersz {row_number}: ilosc musi byc wieksza od zera.")

        item = db.scalar(select(Item).where((Item.sku == code) | (Item.barcode == code)))
        if not item:
            raise WmsError(f"Wiersz {row_number}: nie znaleziono produktu {code}.")

        remaining = quantity
        stock_filters = [Stock.item_id == item.id, Stock.quantity > 0]
        if excluded_locations:
            stock_filters.append(Stock.location.not_in(excluded_locations))
        stocks = list(
            db.scalars(
                select(Stock)
                .where(*stock_filters)
                .order_by(Stock.location)
            )
        )
        for stock in stocks:
            if remaining <= 0:
                break
            reserved_quantity = reserved_by_sku_location.get((item.sku, stock.location), 0)
            available_quantity = max(stock.quantity - reserved_quantity, 0)
            if available_quantity <= 0:
                continue
            pick_quantity = min(remaining, available_quantity)
            db.add(
                PickingTask(
                    batch_id=batch_id,
                    sku=item.sku,
                    source_location=stock.location,
                    target_location=target,
                    quantity=pick_quantity,
                    status="pending",
                )
            )
            reserved_by_sku_location[(item.sku, stock.location)] = reserved_quantity + pick_quantity
            created += 1
            remaining -= pick_quantity
        if remaining > 0:
            db.add(
                PickingTask(
                    batch_id=batch_id,
                    sku=item.sku,
                    source_location=None,
                    target_location=target,
                    quantity=remaining,
                    status="blocked",
                )
            )
            created += 1
            blocked += 1

    if created == 0:
        raise WmsError("Nie znaleziono zadnych pozycji do pickingu.")
    db.commit()
    return created, blocked


def get_reserved_by_sku_location(db: Session) -> dict[tuple[str, str], int]:
    excluded_locations = get_picking_target_locations(db)
    filters = [
        PickingTask.status.in_(RESERVED_PICKING_STATUSES),
        PickingTask.source_location.is_not(None),
    ]
    if excluded_locations:
        filters.append(PickingTask.source_location.not_in(excluded_locations))
    rows = db.execute(
        select(
            PickingTask.sku,
            PickingTask.source_location,
            func.coalesce(func.sum(PickingTask.quantity), 0),
        )
        .where(*filters)
        .group_by(PickingTask.sku, PickingTask.source_location)
    ).all()
    return {(sku, location): int(quantity or 0) for sku, location, quantity in rows if location}


def get_picking_target_locations(db: Session) -> set[str]:
    return {
        location
        for location in db.scalars(select(PickingTask.target_location).where(PickingTask.target_location.is_not(None)))
        if location
    }


def block_invalid_picking_sources(db: Session) -> None:
    excluded_locations = get_picking_target_locations(db)
    if not excluded_locations:
        return
    tasks = list(
        db.scalars(
            select(PickingTask)
            .where(
                PickingTask.status.in_(RESERVED_PICKING_STATUSES),
                PickingTask.source_location.in_(excluded_locations),
            )
            .with_for_update()
        )
    )
    if not tasks:
        return
    for task in tasks:
        task.status = "blocked"
        task.source_location = None
        task.scanner_id = None
        task.operator = None
        task.assigned_at = None
    db.commit()


def get_reserved_by_sku(db: Session) -> dict[str, int]:
    excluded_locations = get_picking_target_locations(db)
    filters = [
        PickingTask.status.in_(RESERVED_PICKING_STATUSES),
        PickingTask.source_location.is_not(None),
    ]
    if excluded_locations:
        filters.append(PickingTask.source_location.not_in(excluded_locations))
    rows = db.execute(
        select(
            PickingTask.sku,
            func.coalesce(func.sum(PickingTask.quantity), 0),
        )
        .where(*filters)
        .group_by(PickingTask.sku)
    ).all()
    return {sku: int(quantity or 0) for sku, quantity in rows}


def get_shipping_picking_batch_statuses(db: Session) -> dict[str, str]:
    rows = db.execute(
        select(
            PickingTask.batch_id,
            func.count(PickingTask.id),
            func.sum(case((PickingTask.status == "done", 1), else_=0)),
            func.sum(case((PickingTask.status == "closed", 1), else_=0)),
        )
        .group_by(PickingTask.batch_id)
    ).all()
    statuses = {}
    for batch_id, total, done, closed in rows:
        total = int(total or 0)
        done = int(done or 0)
        closed = int(closed or 0)
        if total <= 0 or done <= 0:
            continue
        if done >= total:
            statuses[batch_id] = "zebrany"
        elif done + closed >= total:
            statuses[batch_id] = "zebrany czesciowo"
    return statuses


def ensure_allocation_workspace(db: Session, workspace_id: str) -> AllocationWorkspace:
    workspace = db.scalar(select(AllocationWorkspace).where(AllocationWorkspace.workspace_id == workspace_id))
    if not workspace:
        raise HTTPException(status_code=404, detail="Nie znaleziono alokacji roboczej.")
    return workspace


def add_allocation_event(
    db: Session,
    workspace_id: str,
    event_type: str,
    description: str,
    source_filename: str | None = None,
    sku: str | None = None,
    color: str | None = None,
    pallet_count: int | None = None,
    carton_count: int | None = None,
    undo_payload: dict | None = None,
) -> None:
    db.add(
        AllocationEvent(
            workspace_id=workspace_id,
            event_type=event_type,
            description=description[:500],
            source_filename=source_filename,
            sku=sku,
            color=color,
            pallet_count=pallet_count,
            carton_count=carton_count,
            undo_payload=json.dumps(undo_payload) if undo_payload else None,
        )
    )


def parse_count_from_message(message: str, label: str) -> int | None:
    match = re.search(rf"(\d+)\s+{re.escape(label)}", message)
    return int(match.group(1)) if match else None


def parse_carton_count_from_message(message: str) -> int | None:
    return parse_count_from_message(message, "kartonow")


def allocation_pallet_snapshot(pallets: list[DeliveryPallet]) -> list[dict[str, str | None]]:
    return [
        {
            "pallet_code": pallet.pallet_code,
            "allocation_status": pallet.allocation_status,
            "layout_row": pallet.layout_row,
            "layout_position": pallet.layout_position,
        }
        for pallet in pallets
    ]


def restore_allocation_pallet_snapshot(
    db: Session,
    workspace_id: str,
    pallet_snapshot: list[dict],
) -> int:
    restored = 0
    for state in pallet_snapshot:
        pallet_code = str(state.get("pallet_code") or "")
        if not pallet_code:
            continue
        pallet = db.scalar(
            select(DeliveryPallet).where(
                DeliveryPallet.workspace_id == workspace_id,
                DeliveryPallet.pallet_code == pallet_code,
            )
        )
        if not pallet:
            continue
        pallet.allocation_status = state.get("allocation_status") or pallet.allocation_status
        pallet.layout_row = state.get("layout_row")
        pallet.layout_position = state.get("layout_position")
        restored += 1
    return restored


def import_delivery_xlsx(
    db: Session,
    workspace_id: str,
    filename: str,
    content: bytes,
) -> AllocationImportOut:
    if not filename.lower().endswith(".xlsx"):
        raise WmsError("Rozladunek musi byc plikiem XLSX.")
    if not filename.lower().startswith("rozladunek_"):
        raise WmsError("Importowane sa tylko pliki rozladunek_*.xlsx.")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise WmsError("Brakuje biblioteki openpyxl do odczytu XLSX.") from exc

    rows = []
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    header_row = 7
    headers = [normalize_header(cell_to_text(value)) for value in next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    if not any(headers):
        raise WmsError("Nie znaleziono naglowkow rozladunku w wierszu 7.")

    for row_values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = {headers[index]: cell_to_text(value) for index, value in enumerate(row_values) if index < len(headers)}
        sku = row.get("modelokolor", "").strip()
        pallet = row.get("paleta", "").strip()
        if not sku and not pallet:
            continue
        if pallet.upper() == "SUMA":
            continue
        quantity = parse_int(row.get("total box", "0"))
        if not sku or not pallet or quantity <= 0:
            continue
        rows.append(
            {
                "sku": sku,
                "color": row.get("colour", ""),
                "kind": row.get("rodzaj", ""),
                "size": row.get("rozmiar", ""),
                "ean": normalize_code_value(row.get("ean", "")),
                "pallet_no": pallet,
                "quantity": quantity,
                "sumy": parse_optional_int(row.get("sumy", "")),
                "allocation_label": row.get("alokacja", ""),
                "overflow_used": row.get("overflow used", ""),
            }
        )

    if not rows:
        raise WmsError("Nie znaleziono pozycji w pliku rozladunku.")

    old_import = db.scalar(
        select(DeliveryImport).where(
            DeliveryImport.workspace_id == workspace_id,
            DeliveryImport.source_filename == filename[:240],
        )
    )
    replaced = bool(old_import)
    if old_import:
        delete_delivery_import(db, old_import.delivery_id)

    delivery_ref = parse_delivery_ref(filename)
    delivery_id = uuid4().hex[:12].upper()
    db.add(
        DeliveryImport(
            delivery_id=delivery_id,
            workspace_id=workspace_id,
            source_filename=filename[:240],
            delivery_ref=delivery_ref,
            total_cartons=sum(row["quantity"] for row in rows),
        )
    )

    pallet_totals: dict[str, int] = {}
    for row in rows:
        pallet_code = format_pallet_code(delivery_ref or delivery_id, row["pallet_no"])
        pallet_totals[pallet_code] = pallet_totals.get(pallet_code, 0) + row["quantity"]
        db.add(
            DeliveryPalletContent(
                delivery_id=delivery_id,
                workspace_id=workspace_id,
                pallet_code=pallet_code,
                sku=row["sku"],
                color=row["color"] or None,
                kind=row["kind"] or None,
                size=row["size"] or None,
                ean=row["ean"] or None,
                quantity_cartons=row["quantity"],
                sumy=row["sumy"],
                allocation_label=row["allocation_label"] or None,
                overflow_used=row["overflow_used"] or None,
            )
        )

    for pallet_code, total in pallet_totals.items():
        pallet_no = pallet_code.rsplit("-P", 1)[-1].lstrip("0") or pallet_code
        db.add(
            DeliveryPallet(
                delivery_id=delivery_id,
                workspace_id=workspace_id,
                pallet_no=pallet_no,
                pallet_code=pallet_code,
                total_cartons=total,
                allocation_status="roboczo_pod_alokacje",
            )
        )
    db.commit()
    return AllocationImportOut(
        workspace_id=workspace_id,
        source_filename=filename,
        imported=len(rows),
        replaced=replaced,
        message=f"Zaimportowano rozladunek: {len(pallet_totals)} palet, {sum(row['quantity'] for row in rows)} kartonow.",
    )


def import_plan_xlsx(
    db: Session,
    workspace_id: str,
    filename: str,
    content: bytes,
) -> AllocationImportOut:
    if not filename.lower().endswith(".xlsx"):
        raise WmsError("Plan alokacji musi byc plikiem XLSX.")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise WmsError("Brakuje biblioteki openpyxl do odczytu XLSX.") from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    source_filename = filename[:240]
    old_items = list(
        db.scalars(
            select(AllocationPlanItem).where(
                AllocationPlanItem.workspace_id == workspace_id,
                AllocationPlanItem.source_filename == source_filename,
            )
        )
    )
    replaced = bool(old_items)
    for item in old_items:
        db.delete(item)

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise WmsError("Plik alokacji jest pusty.")
    headers = [normalize_header(cell_to_text(value)) for value in rows[0]]
    created = 0
    for row_values in rows[1:]:
        row = {headers[index]: cell_to_text(value) for index, value in enumerate(row_values) if index < len(headers)}
        mdk = row.get("mdk", "").strip()
        ean_prepack = normalize_code_value(row.get("ean prepack", ""))
        if not mdk and not ean_prepack:
            continue
        db.add(
            AllocationPlanItem(
                workspace_id=workspace_id,
                source_filename=source_filename,
                mdk=mdk or ean_prepack,
                season=row.get("sezon", "") or None,
                assortment_group=row.get("grupa asort.", "") or row.get("grupa asort", "") or None,
                model=row.get("model", "") or None,
                color=row.get("kolor", "") or None,
                color_code=row.get("kol", "") or None,
                category=row.get("odziez/ akcesoria", "") or row.get("odziez akcesoria", "") or None,
                supplier=row.get("dostawca", "") or None,
                delivery_plan=row.get("dostawa plan", "") or None,
                ean_prepack=ean_prepack or None,
            )
        )
        created += 1
    if created == 0:
        raise WmsError("Nie znaleziono pozycji w pliku alokacji.")
    db.commit()
    return AllocationImportOut(
        workspace_id=workspace_id,
        source_filename=filename,
        imported=created,
        replaced=replaced,
        message=f"Zaimportowano plan alokacji: {created} pozycji.",
    )


def delete_delivery_import(db: Session, delivery_id: str) -> None:
    for model in (DeliveryPalletContent, DeliveryPallet, DeliveryImport):
        rows = list(db.scalars(select(model).where(model.delivery_id == delivery_id)))
        for row in rows:
            db.delete(row)


def delete_allocation_workspace_rows(db: Session, workspace_id: str) -> None:
    for model in (AllocationPlanItem, DeliveryPalletContent, DeliveryPallet, DeliveryImport, AllocationWorkspace):
        rows = list(db.scalars(select(model).where(model.workspace_id == workspace_id)))
        for row in rows:
            db.delete(row)


def active_pallet_codes(db: Session, workspace_id: str) -> set[str]:
    return set(
        db.scalars(
            select(DeliveryPallet.pallet_code).where(
                DeliveryPallet.workspace_id == workspace_id,
                DeliveryPallet.allocation_status != "usuniete_z_alokacji",
            )
        )
    )


def get_section_pallets(
    db: Session,
    workspace_id: str,
    sku: str,
    color: str | None,
) -> list[DeliveryPallet]:
    content_filters = [
        DeliveryPalletContent.workspace_id == workspace_id,
        DeliveryPalletContent.sku == sku,
    ]
    if color:
        content_filters.append(DeliveryPalletContent.color == color)
    pallet_codes = set(db.scalars(select(DeliveryPalletContent.pallet_code).where(*content_filters)))
    if not pallet_codes:
        return []
    pallets = list(
        db.scalars(
            select(DeliveryPallet)
            .where(
                DeliveryPallet.workspace_id == workspace_id,
                DeliveryPallet.pallet_code.in_(pallet_codes),
                DeliveryPallet.allocation_status != "usuniete_z_alokacji",
            )
            .order_by(DeliveryPallet.pallet_code)
        )
    )
    contents_by_pallet = get_allocation_contents_by_pallet(db, workspace_id)
    for pallet in pallets:
        if not pallet.layout_row:
            contents = contents_by_pallet.get(pallet.pallet_code, [])
            pallet.layout_row = allocation_layout_row([content.kind for content in contents])
    return pallets


def set_section_position(pallets: list[DeliveryPallet], target_position: str) -> None:
    counters: dict[str, int] = {}
    for pallet in sorted(pallets, key=lambda row: (row.layout_row or "", row.pallet_code)):
        row_name = pallet.layout_row or "NIEOKRESLONE"
        counters[row_name] = counters.get(row_name, 0) + 1
        pallet.layout_position = target_position if counters[row_name] == 1 else f"{target_position}.{counters[row_name]}"


def compact_layout_positions(db: Session, workspace_id: str) -> None:
    contents_by_pallet = get_allocation_contents_by_pallet(db, workspace_id)
    auto_layout = build_allocation_layout(db, workspace_id, contents_by_pallet)
    pallets = list(
        db.scalars(
            select(DeliveryPallet)
            .where(
                DeliveryPallet.workspace_id == workspace_id,
                DeliveryPallet.allocation_status != "usuniete_z_alokacji",
            )
            .order_by(DeliveryPallet.pallet_code)
        )
    )
    pallets.sort(
        key=lambda pallet: (
            position_sort_key(pallet.layout_position or auto_layout.get(pallet.pallet_code, {}).get("position")),
            rowSortBackend(pallet.layout_row or auto_layout.get(pallet.pallet_code, {}).get("row")),
            pallet.pallet_code,
        )
    )
    grouped_sections: list[tuple[str, str | None, list[DeliveryPallet]]] = []
    section_by_key: dict[tuple[str, str | None], list[DeliveryPallet]] = {}
    for pallet in pallets:
        contents = contents_by_pallet.get(pallet.pallet_code, [])
        if not pallet.layout_row:
            pallet.layout_row = allocation_layout_row([content.kind for content in contents])
        sku = dominant_value([content.sku for content in contents])
        color = dominant_value([content.color for content in contents])
        key = (sku, color or None)
        if key not in section_by_key:
            section_by_key[key] = []
            grouped_sections.append((sku, color or None, section_by_key[key]))
        section_by_key[key].append(pallet)
    for index, (_sku, _color, section_pallets) in enumerate(grouped_sections, start=1):
        set_section_position(section_pallets, str(index))


def allocation_workspace_out(db: Session, workspace: AllocationWorkspace) -> AllocationWorkspaceOut:
    plan_skus, plan_eans = get_allocation_plan_keys(db, workspace.workspace_id)
    contents_by_pallet = get_allocation_contents_by_pallet(db, workspace.workspace_id)
    contents = [content for rows in contents_by_pallet.values() for content in rows]
    confirmed = sum(
        content.quantity_cartons
        for content in contents
        if allocation_row_status(content.sku, content.ean, plan_skus, plan_eans) == "potwierdzone"
    )
    total = sum(content.quantity_cartons for content in contents)
    total_pallets = db.scalar(
        select(func.count(DeliveryPallet.id)).where(
            DeliveryPallet.workspace_id == workspace.workspace_id,
            DeliveryPallet.allocation_status != "usuniete_z_alokacji",
        )
    )
    plan_items = db.scalar(
        select(func.count(AllocationPlanItem.id)).where(AllocationPlanItem.workspace_id == workspace.workspace_id)
    )
    return AllocationWorkspaceOut(
        workspace_id=workspace.workspace_id,
        name=workspace.name,
        status=workspace.status,
        total_pallets=int(total_pallets or 0),
        total_cartons=total,
        confirmed_cartons=confirmed,
        unconfirmed_cartons=max(total - confirmed, 0) if plan_skus or plan_eans else 0,
        plan_items=int(plan_items or 0),
        created_at=workspace.created_at,
    )


def build_allocation_layout(
    db: Session,
    workspace_id: str,
    contents_by_pallet: dict[str, list[DeliveryPalletContent]],
) -> dict[str, dict[str, str]]:
    plan_order = [
        item.mdk
        for item in db.scalars(
            select(AllocationPlanItem)
            .where(AllocationPlanItem.workspace_id == workspace_id)
            .order_by(AllocationPlanItem.id)
        )
        if item.mdk
    ]
    seen = set()
    ordered_models = []
    for sku in plan_order:
        if sku not in seen:
            ordered_models.append(sku)
            seen.add(sku)

    pallet_model: dict[str, str] = {}
    pallet_row: dict[str, str] = {}
    for pallet_code, contents in contents_by_pallet.items():
        model = dominant_value([content.sku for content in contents])
        row = allocation_layout_row([content.kind for content in contents])
        pallet_model[pallet_code] = model
        pallet_row[pallet_code] = row
        if model and model not in seen:
            ordered_models.append(model)
            seen.add(model)

    position_by_model = {sku: index + 1 for index, sku in enumerate(ordered_models)}
    counters: dict[tuple[str, str], int] = {}
    layout: dict[str, dict[str, str]] = {}
    for pallet_code in sorted(contents_by_pallet):
        model = pallet_model.get(pallet_code, "")
        row = pallet_row.get(pallet_code, "MIESZANE")
        base_position = position_by_model.get(model, 0)
        key = (row, model)
        counters[key] = counters.get(key, 0) + 1
        position = f"{base_position}" if base_position else ""
        if counters[key] > 1:
            position = f"{position}.{counters[key]}" if position else str(counters[key])
        layout[pallet_code] = {
            "row": row,
            "position": position,
        }
    return layout


def allocation_layout_row(kinds: list[str | None]) -> str:
    normalized = {normalize_header(kind or "") for kind in kinds if kind}
    has_prepack = any("prepak" in kind for kind in normalized)
    has_luz = any("luz" in kind for kind in normalized)
    if has_prepack and not has_luz:
        return "PREPAK"
    if has_luz and not has_prepack:
        return "LUZ"
    if has_prepack and has_luz:
        return "MIESZANE"
    return "NIEOKRESLONE"


def dominant_value(values: list[str | None]) -> str:
    counts: dict[str, int] = {}
    for value in values:
        if not value:
            continue
        counts[value] = counts.get(value, 0) + 1
    if not counts:
        return ""
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[0][0]


def position_sort_key(value: str | None) -> tuple[int, ...]:
    if not value:
        return (9999,)
    parts = []
    for part in str(value).split("."):
        try:
            parts.append(int(float(part)))
        except ValueError:
            parts.append(9999)
    return tuple(parts)


def rowSortBackend(row: str | None) -> int:
    return {"PREPAK": 1, "LUZ": 2, "MIESZANE": 3, "NIEOKRESLONE": 4}.get(row or "", 5)


def get_delivery_imports_by_id(db: Session, workspace_id: str) -> dict[str, DeliveryImport]:
    return {
        row.delivery_id: row
        for row in db.scalars(select(DeliveryImport).where(DeliveryImport.workspace_id == workspace_id))
    }


def get_allocation_contents_by_pallet(
    db: Session,
    workspace_id: str,
    include_removed: bool = False,
) -> dict[str, list[DeliveryPalletContent]]:
    pallet_codes = active_pallet_codes(db, workspace_id) if not include_removed else None
    contents_by_pallet: dict[str, list[DeliveryPalletContent]] = {}
    query = select(DeliveryPalletContent).where(DeliveryPalletContent.workspace_id == workspace_id)
    if pallet_codes is not None:
        if not pallet_codes:
            return {}
        query = query.where(DeliveryPalletContent.pallet_code.in_(pallet_codes))
    for content in db.scalars(query):
        contents_by_pallet.setdefault(content.pallet_code, []).append(content)
    return contents_by_pallet


def get_allocation_plan_keys(db: Session, workspace_id: str) -> tuple[set[str], set[str]]:
    items = list(db.scalars(select(AllocationPlanItem).where(AllocationPlanItem.workspace_id == workspace_id)))
    skus = {item.mdk for item in items if item.mdk}
    eans = {part for item in items for part in split_codes(item.ean_prepack)}
    return skus, eans


def get_allocation_content_keys(db: Session, workspace_id: str) -> tuple[set[str], set[str]]:
    contents_by_pallet = get_allocation_contents_by_pallet(db, workspace_id)
    contents = [content for rows in contents_by_pallet.values() for content in rows]
    skus = {content.sku for content in contents if content.sku}
    eans = {part for content in contents for part in split_codes(content.ean)}
    return skus, eans


def allocation_content_status(
    contents: list[DeliveryPalletContent],
    plan_skus: set[str],
    plan_eans: set[str],
) -> str:
    if not plan_skus and not plan_eans:
        return "roboczo"
    statuses = {allocation_row_status(content.sku, content.ean, plan_skus, plan_eans) for content in contents}
    if statuses == {"potwierdzone"}:
        return "potwierdzone"
    if "potwierdzone" in statuses:
        return "czesciowo potwierdzone"
    return "niepotwierdzone"


def allocation_row_status(sku: str | None, ean: str | None, plan_skus: set[str], plan_eans: set[str]) -> str:
    if not plan_skus and not plan_eans:
        return "roboczo"
    if sku and sku in plan_skus:
        return "potwierdzone"
    if any(code in plan_eans for code in split_codes(ean)):
        return "potwierdzone"
    return "niepotwierdzone"


def allocation_plan_item_found(item: AllocationPlanItem, content_skus: set[str], content_eans: set[str]) -> bool:
    if item.mdk and item.mdk in content_skus:
        return True
    return any(code in content_eans for code in split_codes(item.ean_prepack))


def split_codes(value: str | None) -> set[str]:
    if not value:
        return set()
    return {part.strip() for part in re.split(r"[/,; ]+", value) if part.strip()}


def parse_delivery_ref(filename: str) -> str | None:
    match = re.search(r"_(\d+)(?:\.[^.]+)?$", filename)
    return match.group(1) if match else None


def format_pallet_code(delivery_ref: str, pallet_no: str) -> str:
    if re.fullmatch(r"\d+(\.0)?", pallet_no):
        number = int(float(pallet_no))
        return f"{delivery_ref}-P{number:03d}"
    return f"{delivery_ref}-P{pallet_no}"


def parse_int(value: object) -> int:
    text_value = cell_to_text(value).replace(",", ".")
    try:
        return int(float(text_value))
    except ValueError:
        return 0


def parse_optional_int(value: object) -> int | None:
    parsed = parse_int(value)
    return parsed if parsed else None


def normalize_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def normalize_code_value(value: str) -> str:
    value = value.strip()
    if re.fullmatch(r"\d+\.0", value):
        return value[:-2]
    return value


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def get_picking_value(row: dict[str, str], field: str) -> str:
    aliases = PICKING_HEADER_ALIASES[field]
    for alias in aliases:
        value = row.get(normalize_header(alias))
        if value:
            return value
    return ""


def picking_batch_out(
    batch_id: str,
    source_filename: str | None,
    total_tasks: int,
    assigned_tasks: int,
    done_tasks: int,
    active_tasks: int,
    canceled_tasks: int,
    closed_tasks: int,
    created_at: object,
) -> PickingBatchOut:
    if total_tasks <= 0:
        status = "oczekuje"
        progress_percent = 0
    elif done_tasks >= total_tasks:
        status = "zebrany"
        progress_percent = 100
    elif done_tasks + closed_tasks >= total_tasks and closed_tasks > 0:
        status = "zebrany czesciowo"
        progress_percent = round(done_tasks * 100 / total_tasks)
    elif done_tasks + canceled_tasks >= total_tasks:
        status = "anulowany"
        progress_percent = round(done_tasks * 100 / total_tasks)
    elif done_tasks > 0 or active_tasks > 0:
        status = "w trakcie"
        progress_percent = round(done_tasks * 100 / total_tasks)
    else:
        status = "oczekuje"
        progress_percent = 0
    return PickingBatchOut(
        batch_id=batch_id,
        source_filename=source_filename,
        total_tasks=total_tasks,
        assigned_tasks=assigned_tasks,
        status=status,
        progress_percent=progress_percent,
        created_at=created_at,
    )


def picking_task_out(db: Session, task: PickingTask) -> PickingTaskOut:
    item = db.scalar(select(Item).where(Item.sku == task.sku))
    return PickingTaskOut(
        id=task.id,
        batch_id=task.batch_id,
        sku=task.sku,
        barcode=item.barcode if item else None,
        name=item.name if item else None,
        source_location=task.source_location,
        target_location=task.target_location,
        quantity=task.quantity,
        status=task.status,
        scanner_id=task.scanner_id,
        operator=task.operator,
        assigned_at=task.assigned_at,
        picked_at=task.picked_at,
        created_at=task.created_at,
    )


def shipping_out(db: Session, task: PickingTask, batch: PickingBatch | None, picking_status: str) -> ShippingOut:
    item = db.scalar(select(Item).where(Item.sku == task.sku))
    return ShippingOut(
        scan_at=task.picked_at,
        batch_id=task.batch_id,
        source_filename=batch.source_filename if batch else None,
        picking_status=picking_status,
        shipping_status="gotowe do wysylki",
        sku=task.sku,
        barcode=item.barcode if item else None,
        name=item.name if item else None,
        source_location=task.source_location,
        target_location=task.target_location,
        quantity=task.quantity,
        operator=task.operator,
        scanner_id=task.scanner_id,
        assigned_at=task.assigned_at,
        picked_at=task.picked_at,
        created_at=task.created_at,
    )
